"""Configuration globale, accès base de données et classement des barrages."""
import os
import sqlite3
import sys
import pandas as pd
import streamlit as st
from typing import List, Optional

from src.models import DamScore

if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "dams.db")

# Path to current study Excel file
CURRENT_STUDY_EXCEL = os.path.join(BASE_DIR, "..", "fichier excel calcul evaporation (2).xlsx")

# Tarif STEG 2025 obligatoire
J1_STEG_2025 = 0.307

@st.cache_resource
def get_connection():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def load_constants(conn) -> dict:
    df = pd.read_sql("SELECT * FROM constants", conn)
    constants = {row['key']: row['value'] for _, row in df.iterrows()}
    constants['J1'] = J1_STEG_2025
    return constants

def update_j1_in_db(conn) -> None:
    conn.execute("UPDATE constants SET value=? WHERE key='J1'", (J1_STEG_2025,))
    conn.commit()

def load_dams(conn) -> pd.DataFrame:
    return pd.read_sql("SELECT * FROM dams ORDER BY id", conn)

def load_evap(conn, dam_id: int) -> pd.DataFrame:
    return pd.read_sql(
        "SELECT * FROM evaporation_monthly WHERE dam_id=? ORDER BY month",
        conn, params=(dam_id,)
    )

def load_thermal(conn, dam_id: int) -> pd.DataFrame:
    return pd.read_sql(
        "SELECT * FROM thermal_monthly WHERE dam_id=? ORDER BY month",
        conn, params=(dam_id,)
    )

def load_scenarios(conn) -> pd.DataFrame:
    return pd.read_sql("SELECT * FROM economic_scenarios ORDER BY id", conn)

def get_aquatic_gain(dam_id: int) -> dict:
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT gain_percent, gain_kwh FROM aquatic_gain WHERE dam_id=?", (dam_id,))
    row = c.fetchone()
    if row:
        return {"gain_percent": row[0], "gain_kwh": row[1]}
    return {"gain_percent": 0.0, "gain_kwh": 0}

def load_dam_evaporation_from_excel(dam_name: str) -> dict:
    """Load dam evaporation totals from current study Excel file (fichier excel calcul evaporation (2))."""
    excel_path = os.path.join(os.path.dirname(__file__), "..", "fichier excel calcul evaporation (2).xlsx")
    
    if not os.path.exists(excel_path):
        # Return empty to use DB values as fallback
        return {}
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        
        sheet_map = {
            "Sidi Saad": "Barrage sidi saad ",
            "Sidi Salem": "Sidi salem ",
            "Sidi El Barrak": "sidi Barrak",
            "Bouhertma": "Bouhertma",
            "Sejnane": "Sejnane",
        }
        
        sheet_name = sheet_map.get(dam_name)
        if not sheet_name or sheet_name not in wb.sheetnames:
            return {}
        
        ws = wb[sheet_name]
        
        # Look for "Total économies" in the sheet
        total_saved = None
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val and 'Total économies' in str(val):
                    # Look for number in next cells
                    idx = list(row).index(val)
                    for j in range(idx + 1, min(idx + 4, len(row))):
                        next_val = row[j]
                        if next_val:
                            num_str = str(next_val).replace(' ', '').replace('m³', '').replace(',', '.')
                            try:
                                total_saved = float(num_str)
                            except:
                                pass
                    break
            if total_saved is not None:
                break
        
        wb.close()
        
        if total_saved is None:
            return {}
        
        # The totals in Excel are for 20 MWc - convert to per MWc
        return {"economie_m3_per_mwc": total_saved / 20.0}
    except ImportError:
        # openpyxl not available - use DB values as fallback
        return {}
    except Exception as e:
        # Other errors - use DB values as fallback
        return {}

# ----------------------------------------------------------------------
# Classement prioritaire (rapport modifications)
# ----------------------------------------------------------------------
def _score_color(score: float) -> str:
    if score >= 80:
        return "green"
    elif score >= 65:
        return "lightgreen"
    elif score >= 50:
        return "yellow"
    elif score >= 35:
        return "orange"
    else:
        return "red"

def _normalize(series: pd.Series) -> pd.Series:
    mn, mx = series.min(), series.max()
    if mx == mn:
        return pd.Series([50.0] * len(series), index=series.index)
    return (series - mn) / (mx - mn) * 100

def compute_dam_scores(conn, power_mwc: float = 20.0, dam_totals: dict = None) -> List[DamScore]:
    """
    Compute dam scores based on multi-criteria analysis.
    If dam_totals is provided (from Excel), use those instead of database values.
    """
    # Hardcoded dam productible values (PV characteristics)
    dam_productible = {
        1: 1703,  # Sidi Salem
        2: 1646,  # Sidi El Barrak
        3: 1696,  # Bouhertma
        4: 1660,  # Sejnane
        5: 1780,  # Sidi Saad
    }
    
    # Aquatic gains per dam (average gain_percent from monthly values)
    dam_aquatic_gain = {
        1: 6.15,  # Sidi Salem
        2: 4.82,  # Sidi El Barrak
        3: 4.78,  # Bouhertma
        4: 4.75,  # Sejnane
        5: 4.56,  # Sidi Saad
    }
    
    # Dam constraints (name, is_ramsar, max_coverage_rate)
    dam_constraints = {
        1: ("Sidi Salem", False, 15.0),
        2: ("Sidi El Barrak", False, 15.0),
        3: ("Bouhertma", False, 15.0),
        4: ("Sejnane", False, 5.0),  # AEP, not Ramsar
        5: ("Sidi Saad", True, 5.0),  # Ramsar site
    }
    
    rows = []
    for dam_id in sorted(dam_constraints.keys()):
        dam_name, is_ramsar, _ = dam_constraints[dam_id]
        
        # Use Excel totals if provided, otherwise fallback to DB
        if dam_totals and dam_name in dam_totals:
            water_per_mwc = dam_totals[dam_name].get("economie_m3_per_mwc", 0)
        else:
            # Fallback to DB values
            conn_local = conn or get_connection()
            dam = load_dams(conn_local)[load_dams(conn_local)['id'] == dam_id].iloc[0]
            water_per_mwc = float(dam['economy_water_m3_mwc'])
        
        aquatic_pct = dam_aquatic_gain[dam_id]
        
        production_gwh = dam_productible[dam_id] * power_mwc / 1000.0
        water_m3 = water_per_mwc * power_mwc
        pr_raw = dam_productible[dam_id]
        
        rows.append({
            'dam_id': dam_id,
            'dam_name': dam_name,
            'production_gwh': production_gwh,
            'water_m3': water_m3,
            'aquatic_pct': aquatic_pct,
            'pr_raw': pr_raw,
            'has_ramsar': is_ramsar,
        })
    
    df = pd.DataFrame(rows)
    df['prod_norm'] = _normalize(df['production_gwh'])
    df['water_norm'] = _normalize(df['water_m3'])
    df['aquatic_norm'] = _normalize(df['aquatic_pct'])
    df['pr_norm'] = _normalize(df['pr_raw'])
    df['constraint_norm'] = df['has_ramsar'].apply(lambda r: 0.0 if r else 100.0)

    df['score'] = (df['prod_norm'] * 0.30 +
                   df['water_norm'] * 0.25 +
                   df['aquatic_norm'] * 0.20 +
                   df['pr_norm'] * 0.15 +
                   df['constraint_norm'] * 0.10)

    df = df.sort_values('score', ascending=False).reset_index(drop=True)

    results = []
    for rank, (_, row) in enumerate(df.iterrows(), start=1):
        results.append(DamScore(
            dam_id=int(row['dam_id']),
            dam_name=row['dam_name'],
            score=round(row['score'], 1),
            rank=rank,
            color=_score_color(row['score']),
            production_score=round(row['prod_norm'], 1),
            water_score=round(row['water_norm'], 1),
            aquatic_score=round(row['aquatic_norm'], 1),
            pr_score=round(row['pr_norm'], 1),
            constraint_score=round(row['constraint_norm'], 1),
            production_gwh=round(row['production_gwh'], 3),
            water_m3=round(row['water_m3'], 0),
            aquatic_gain_pct=round(row['aquatic_pct'], 2),
            pr_pct=round(row['pr_raw'] * 100, 1),
            has_ramsar=bool(row['has_ramsar']),
        ))
    return results

def scores_to_dataframe(scores: List[DamScore]) -> pd.DataFrame:
    return pd.DataFrame([{
        'Rang': f"{s.color} #{s.rank}",
        'Barrage': s.dam_name,
        'Score': s.score,
        'Production (GWh/an)': s.production_gwh,
        'Eau (m³/an)': int(s.water_m3),
        'Gain aquatique (%)': s.aquatic_gain_pct,
        'Ramsar': '⚠️ Oui' if s.has_ramsar else '✅ Non',
    } for s in scores])